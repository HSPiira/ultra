"""XLSX export strategy implementation."""

from io import BytesIO
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .base import ExportStrategy


class XLSXExporter(ExportStrategy):
    """
    Concrete strategy for Excel (XLSX) export.

    Follows Single Responsibility: Only handles XLSX format generation.
    Includes professional styling with headers, auto-sizing, and formatting.
    """

    def __init__(self, sheet_name: str = "Export"):
        """
        Initialize XLSX exporter.

        Args:
            sheet_name: Name for the Excel worksheet (default: "Export")
        """
        self.sheet_name = sheet_name

    def export(self, headers: List[str], rows: List[List[Any]]) -> bytes:
        """
        Export data to XLSX format with professional styling.

        Args:
            headers: Column headers
            rows: Data rows

        Returns:
            XLSX content as bytes
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.sheet_name

        # Style header row
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="left", vertical="center")

        # Write headers with styling
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        for row in rows:
            # Convert all values to strings, handle None
            cleaned_row = [
                str(value) if value is not None else ""
                for value in row
            ]
            worksheet.append(cleaned_row)

        # Auto-fit column widths (with sensible maximum)
        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = 0
            for cell in column_cells:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            # Set width with maximum of 50 characters
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def get_content_type(self) -> str:
        """Get MIME type for XLSX."""
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def get_file_extension(self) -> str:
        """Get file extension for XLSX."""
        return 'xlsx'
